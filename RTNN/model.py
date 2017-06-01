"""

"""
import math
import operator
import random
from time import time
from openpyxl import Workbook

from nltk.corpus import stopwords
import numpy as np

from RTNN.tree import Node, TreeNet
from six.moves import cPickle


class Model:

    # region init, save, load
    def __init__(self, dim, alpha):
        """

        :param trainf
        :param devf
        :param dim: 
        :param alpha: 
        """
        self.wb = None

        self.dim = dim
        self.alpha = alpha

        # params
        self.word_map = dict()
        self.L = 0.01 * np.random.randn(self.dim,)

        self.V = 0.01 * np.random.randn(self.dim, 2 * self.dim, 2 * self.dim)
        self.W = 0.01 * np.random.randn(self.dim, 2 * self.dim)
        self.b = 0.01 * np.random.randn(self.dim,)

        self.Ws = 0.01 * np.random.randn(self.dim, self.dim)
        self.bs = 0.01 * np.random.randn(self.dim,)

        self.paramsShape = [self.V.shape, self.W.shape, self.b.shape, self.Ws.shape, self.bs.shape]

        # gradient
        self.dL, self.dV, self.dW, self.db, self.dWs, self.dbs = self.default_grad()

        # sum gradient ** 2
        self.sumdL2, self.sumdV2, self.sumdW2, self.sumdb2, self.sumdWs2, self.sumdbs2 = self.default_grad()

    def default_grad(self):
        return [dict()] + [np.zeros(p) for p in self.paramsShape]

    def save(self, file):
        """
        
        :param file: 
        :return: 
        """
        with open(file, 'wb') as f:
            cPickle.dump(self, f)
        print("save successfully")

    @staticmethod
    def load(file):
        """
        
        :param file: 
        :return: 
        """
        with open(file, 'rb') as f:
            return cPickle.load(f)

    # endregion

    # region activate function
    def activate_node(self, node, left, right):
        """
        
        :param node: 
        :param left: 
        :param right: 
        :return: 
        """
        node.actf = Model.activate_function(self.V, self.W, self.b, left, right)
        node.prob = Model.softmax(self.Ws, self.bs, node.actf)

    @staticmethod
    def activate_function(tensor, weight, bias, left, right):
        """
        
        :param tensor: 
        :param weight: 
        :param bias: 
        :param left: 
        :param right: 
        :return: 
        """
        lr = np.concatenate((left, right))
        return np.tanh(Model.tensordot(lr, tensor) + np.dot(weight, lr) + bias)

    @staticmethod
    def tensordot(a, b):
        """
        
        :param a: 
        :param b: 
        :return: 
        """
        left = np.asarray([np.dot(a.T, r).tolist() for r in b])
        return np.asarray([np.dot(r, a).tolist() for r in left])

    @staticmethod
    def softmax(weight, bias, x):
        """
        
        :param weight: 
        :param bias: 
        :param x: 
        :return: 
        """
        r = np.dot(weight, x) + bias
        return np.exp(r) / np.sum(np.exp(r), axis=0)

    # endregion

    # region preproccess

    def preproccess(self, trainf, devf, word_limits):
        """
        
        :param trainf: 
        :param devf: 
        :param word_limits: 
        :return: 
        """
        start = time()

        train_trees = TreeNet.load_trees(trainf + ".txt")
        dictionary, count, words_count = Model.get_word_by_tree(train_trees)
        word_ig = Model.compute_info_gain(words_count, count)
        dictionary = Model.extract_dict(word_ig, word_limits)
        self.word_map.update({dictionary[i]: i for i in range(len(dictionary))})
        for i in range(len(self.word_map)):
            self.L = np.vstack((self.L, 0.01 * np.random.randn(self.dim, )))
        with open(trainf + ".p", 'wb') as f:
            for tree in train_trees:
                TreeNet.BFS_traverse(tree.root, self.process_node)
                cPickle.dump(tree, f)

        dev_trees = TreeNet.load_trees(devf + ".txt")
        with open(devf + ".p", 'wb') as f:
            for tree in dev_trees:
                TreeNet.BFS_traverse(tree.root, self.process_node)
                cPickle.dump(tree, f)

        print('Finish proccess data in {t} seconds.'.format(t=time() - start))

    @staticmethod
    def get_word_by_tree(trees):
        dictionary = set()
        count = np.zeros(5,)
        words_count = dict()
        for tree in trees:
            d, c, wc = Model.get_word_by_node(tree.root)
            dictionary |= d
            count += c
            for word in wc.keys():
                try:
                    words_count[word] += wc[word]
                except KeyError:
                    words_count[word] = wc[word]
        return dictionary, count, words_count

    @staticmethod
    def get_word_by_node(node):
        label = int(node.label)
        if node.isLeaf:
            if node.word in stopwords.words("english"):
                return set(), np.zeros(5,), dict()
            dictionary = {node.word}
            count = np.zeros(5,)
            count[label] += 1
            words_count = {node.word: count}
        else:
            left = Model.get_word_by_node(node.left)
            right = Model.get_word_by_node(node.right)
            dictionary = left[0] | right[0]

            count = left[1] + right[1]
            count[label] += 1

            words_count = dict()
            for word in set(left[2]).intersection(set(right[2])):
                words_count[word] = left[2][word] + right[2][word]
                words_count[word][label] += 1
            for word in set(left[2]).difference(set(right[2])):
                words_count[word] = left[2][word]
                words_count[word][label] += 1
            for word in set(right[2]).difference(set(left[2])):
                words_count[word] = right[2][word]
                words_count[word][label] += 1
        return dictionary, count, words_count

    def process_node(self, node):
        """
        
        :param node: 
        :return: 
        """
        if node.isLeaf:
            node.word = self.word_map.get(node.word, 0)
        node.label = int(node.label)

    @staticmethod
    def compute_info_gain(word_count, cc):
        """
        
        :param word_count: 
        :param cc: 
        :return: 
        """
        word_IG = dict()
        dim = len(cc)
        esp = 1e-4

        def f(x):
            return math.log2(x + esp)

        N = sum(cc)
        for word, c in word_count.items():
            Nw = sum(c)
            word_IG[word] = sum([(-cc[i] / N) * f(cc[i] / N)
                                 + (Nw / N) * (c[i] / Nw) * f(c[i] / Nw)
                                 + (1 - Nw / N) * ((cc[i] - c[i]) / (N - Nw + esp)) * f((cc[i] - c[i]) / (N - Nw + esp))
                                 for i in range(dim)])
        return word_IG

    @staticmethod
    def extract_dict(word_ig, limit):
        """
        
        :param word_ig: 
        :param limit: 
        :return: 
        """
        sorted_words = sorted(word_ig.items(), key=operator.itemgetter(1))
        sorted_words.reverse()
        return [word for word, _ in sorted_words[0:limit]]

    def create_rpfile(self, filename, trainf, devf):
        self.wb = Workbook()

        # ==============================================================================================================
        ws = self.wb.active
        ws.title = "data decription"

        ws["A3"] = "Root"
        ws["A2"] = "label"

        ws.merge_cells("B1:F1")
        ws["B1"] = "Train"
        ws["B2"], ws["C2"], ws["D2"], ws["E2"], ws["F2"] = 0, 1, 2, 3, 4
        self.inspect_dataset(trainf + ".p", 3, 2)

        ws.merge_cells("G1:K1")
        ws["G1"] = "Dev"
        ws["G2"], ws["H2"], ws["I2"], ws["J2"], ws["K2"] = 0, 1, 2, 3, 4
        self.inspect_dataset(devf + ".p", 3, 7)

        # ==============================================================================================================
        self.wb.create_sheet("dictionary")
        ws = self.wb["dictionary"]
        for word, index in self.word_map.items():
            ws.cell(row=1+int(index), column=2).value = word

        # ==============================================================================================================
        self.wb.create_sheet("accuracy all")
        ws = self.wb["accuracy all"]
        ws["A2"] = "epoch"
        ws.merge_cells("B1:D1")
        ws["B1"] = "train"
        ws.merge_cells("E1:G1")
        ws["E1"] = "dev"
        ws['B2'], ws['C2'], ws['D2'] = ws['E2'], ws['F2'], ws['G2'] = 'error', 'all', 'root'

        # ==============================================================================================================
        self.wb.create_sheet("accuracy n-gram on train set")
        ws = self.wb["accuracy n-gram on train set"]
        ws['A1'] = 'n-gram'

        # ==============================================================================================================
        self.wb.create_sheet("accuracy n-gram on dev set")
        ws = self.wb["accuracy n-gram on dev set"]
        ws['A1'] = 'n-gram'

        self.wb.save(filename)

    def inspect_dataset(self, filename, i, j):
        with open(filename, 'rb') as file:
            max_ngram = 0
            while True:
                try:
                    tree = cPickle.load(file)
                    num_word = int(tree.root.numWord)
                    max_ngram = num_word if num_word > max_ngram else max_ngram
                    ws = self.wb["data decription"]
                    if ws.cell(row=3, column=j+int(tree.root.label)).value is not None:
                        ws.cell(row=3, column=j+int(tree.root.label)).value += 1
                    else:
                        ws.cell(row=3, column=j + int(tree.root.label)).value = 1
                    self.inspect_node(tree.root, i, j)
                except EOFError:
                    break

    def inspect_node(self, node, i, j):
        ws = self.wb["data decription"]
        if ws.cell(row=i+int(node.numWord), column=1).value is None:
            ws.cell(row=i+int(node.numWord), column=1).value = int(node.numWord)
        if ws.cell(row=i+int(node.numWord), column=j+int(node.label)).value is not None:
            ws.cell(row=i+int(node.numWord), column=j+int(node.label)).value += 1
        else:
            ws.cell(row=i + int(node.numWord), column=j + int(node.label)).value = 1
        if node.isLeaf:
            return
        else:
            self.inspect_node(node.left, i, j)
            self.inspect_node(node.right, i, j)
    # endregion

    # region train_s

    def train(self, trainf, devf, rpf, savef, save_fre, word_limit, epoch, batch_size, step_size=0.01, fudge=1e-8):
        """

        :param trainf:
        :param devf:
        :param rpf:
        :param savef:
        :param save_fre:
        :param word_limit:
        :param epoch:
        :param batch_size:
        :param step_size:
        :param fudge:
        :return:
        """
        self.preproccess(trainf, devf, word_limit)
        self.create_rpfile(rpf, trainf, devf)
        start = time()
        for i in range(epoch):
            self.wb["accuracy all"].cell(row=3 + i, column=1).value = i + 1
            self.wb["accuracy n-gram on train set"].cell(row=1, column=2 + i).value = i + 1
            self.wb["accuracy n-gram on dev set"].cell(row=1, column=2 + i).value = i + 1
            if not i + 1 % save_fre:
                print("Starting epoch {n}...".format(n=i + 1), end='')
            self.sumdL2, self.sumdV2, self.sumdW2, self.sumdb2, self.sumdWs2, self.sumdbs2 = self.default_grad()
            total_tree = 0
            with open(trainf + '.p', 'rb') as f:
                res = np.zeros((4,))
                data = list()
                for j in range(2 * batch_size):
                    data.append(cPickle.load(f))
                batch = list()
                while True:
                    for j in range(batch_size):
                        try:
                            data.append(cPickle.load(f))
                        except EOFError:
                            break
                    if not len(data):
                        break
                    else:
                        random.shuffle(data)
                        for k in range(min(batch_size, len(data))):
                            batch.append(data.pop(0))
                        res += self.train_batch(i, batch, step_size, fudge)
                        total_tree += len(batch)
                        batch.clear()
            self.wb["accuracy all"].cell(row=3 + i, column=2).value = res[0]
            self.wb["accuracy all"].cell(row=3 + i, column=3).value = res[1] * 100 / res[2]
            self.wb["accuracy all"].cell(row=3 + i, column=4).value = res[3] * 100 / total_tree
            if not i + 1 % save_fre:
                self.save(savef)

            res = self.test_dev(devf, i)
            self.wb["accuracy all"].cell(row=3 + i, column=5).value = res[0]
            self.wb["accuracy all"].cell(row=3 + i, column=6).value = res[1] * 100 / res[2]
            self.wb["accuracy all"].cell(row=3 + i, column=7).value = res[3] * 100 / res[4]

        t = time() - start
        print("Finish training in {t}.".format(t=t))
        self.wb.save(rpf)

    def test_dev(self, devf, epoch):
        with open(devf + '.p', 'rb') as file:
            res = np.zeros((3,))
            correct_tree = 0
            total_tree = 0
            while True:
                try:
                    tree = cPickle.load(file)
                    total_tree += 1
                    res += self.forward_prob(tree.root, epoch, mode="dev")
                    if np.argmax(tree.root.prob) == tree.root.label:
                        correct_tree += 1
                except EOFError:
                    break
            res[0] += self.alpha / 2 * (np.sum(self.V ** 2)
                                        + np.sum(self.W ** 2)
                                        + np.sum(self.Ws ** 2)
                                        + np.sum(self.b ** 2)
                                        + np.sum(self.bs ** 2))
            return np.hstack((res, correct_tree, total_tree))

    def train_batch(self, epoch, batch, step_size, fudge_factor):
        """

        :param epoch
        :param batch: 
        :param step_size: 
        :param fudge_factor: 
        :return: 
        """
        res_batch = np.zeros((3,))
        correct_tree = 0
        self.dL, self.dV, self.dW, self.db, self.dWs, self.dbs = self.default_grad()

        for tree in batch:
            res_batch += self.forward_prob(tree.root, epoch)
            if np.argmax(tree.root.prob) == tree.root.label:
                correct_tree += 1
        res_batch[0] += self.alpha / 2 * (np.sum(self.V ** 2)
                                          + np.sum(self.W ** 2)
                                          + np.sum(self.Ws ** 2)
                                          + np.sum(self.b ** 2)
                                          + np.sum(self.bs ** 2))

        for tree in batch:
            self.backward_prob(tree.root)

        scale = 1.0 / len(batch)
        for val in self.dL.values():
            val *= scale

        self.dV = scale * (self.dV + self.alpha * self.V)
        self.dW = scale * (self.dW + self.alpha * self.W)
        self.db = scale * (self.db + self.alpha * self.b)

        self.dWs = scale * (self.dWs + self.alpha * self.Ws)
        self.dbs = scale * (self.dbs + self.alpha * self.bs)

        self.update(step_size, fudge_factor)
        return np.hstack((res_batch, correct_tree))

    def forward_prob(self, node: Node, epoch, mode="train"):
        """
        
        :param node:
        :param epoch
        :param mode:
        :return: 
        """
        err = np.zeros((3,))
        if node.isLeaf:
            node.actf = self.L[node.word]
            node.prob = Model.softmax(self.Ws, self.bs, node.actf)
        else:
            err += self.forward_prob(node.left, epoch, mode)
            err += self.forward_prob(node.right, epoch, mode)
            self.activate_node(node, node.left.actf, node.right.actf)

        if mode == "train":
            ws = self.wb["accuracy n-gram on train set"]
        else:
            ws = self.wb["accuracy n-gram on dev set"]
        if ws.cell(row=1 + int(node.numWord), column=1).value is None:
            ws.cell(row=1 + int(node.numWord), column=1).value = int(node.numWord)
        if ws.cell(row=1 + int(node.numWord), column=2 + epoch).value is None:
            ws.cell(row=1 + int(node.numWord), column=2 + epoch).value = 0
        if np.argmax(node.prob) == node.label:
            ws.cell(row=1 + int(node.numWord), column=2 + epoch).value += 1
        return err \
            + np.asarray([-math.log(node.prob[node.label]), np.argmax(node.prob) == node.label, 1])

    def backward_prob(self, node: Node, err=None):
        """
        
        :param node: 
        :param err: 
        :return: 
        """
        delta = node.prob
        delta[node.label] -= 1.0

        # dWs
        self.dWs += np.outer(delta, node.actf)
        self.dbs += delta

        softmax_err = np.dot(self.Ws.T, delta) * (1 - node.actf ** 2)

        if err is None:
            com_err = softmax_err
        else:
            com_err = softmax_err + err * (1 - node.actf ** 2)

        if node.isLeaf:
            if node.word in self.dL.keys():
                self.dL[node.word] += com_err
            else:
                self.dL[node.word] = com_err
            return

        lr = np.concatenate((node.left.actf, node.right.actf))

        # dV
        self.dV += (np.outer(lr, lr)[..., None] * com_err).T
        # dW
        self.dW += np.outer(com_err, lr)
        # db
        self.db += com_err

        down_err = np.dot(self.W.T, com_err) \
            + np.tensordot(self.V.transpose((0, 2, 1)) + self.V, np.outer(com_err, lr).T, axes=([1, 0], [0, 1]))
        self.backward_prob(node.left, down_err[:self.dim])
        self.backward_prob(node.right, down_err[self.dim:])

    def update(self, step_size, fudge_factor):
        """
        
        :param step_size: 
        :param fudge_factor: 
        :return: 
        """
        for key, val in self.dL.items():
            if key in self.sumdL2.keys():
                self.sumdL2[key] += val ** 2
            else:
                self.sumdL2[key] = val ** 2
            self.L[key] -= step_size / (fudge_factor + np.sqrt(self.sumdL2[key])) * val

        self.sumdV2 += self.dV ** 2
        self.V -= step_size / (fudge_factor + np.sqrt(self.sumdV2)) * self.dV

        self.sumdW2 += self.dW ** 2
        self.W -= step_size / (fudge_factor + np.sqrt(self.sumdW2)) * self.dW

        self.sumdWs2 += self.dWs ** 2
        self.Ws -= step_size / (fudge_factor + np.sqrt(self.sumdWs2)) * self.dWs

        self.sumdb2 += self.db ** 2
        self.b -= step_size / (fudge_factor + np.sqrt(self.sumdb2)) * self.db

        self.sumdbs2 += self.dbs ** 2
        self.bs -= step_size / (fudge_factor + np.sqrt(self.sumdbs2)) * self.dbs

    # endregion

    # region test
    def test(self, testf, rpf):
        """
        
        :param testf: 
        :param rpf: 
        :return: 
        """
        with open(rpf, 'w') as report:
            with open(testf, 'r') as f:
                result = np.zeros((4,))
                for line in f:
                    tree = TreeNet(line)
                    res = self.estimate_sentis(tree)
                    print("correct node: {cn} - total node: {ct} - correct tree: {ctr}- total tree: {ttr}."
                          .format(cn=res[0], ct=res[1], ctr=res[2], ttr=res[3]))
                    result += res
                print("=" * 80)
                report.write("correct node: {cn} - total node: {ct} - correct tree: {ctr}- total tree: {ttr}."
                             .format(cn=result[0], ct=result[1], ctr=result[2], ttr=result[3]))
                return result

    def estimate_sentis(self, tree):
        """
        
        :param tree: 
        :return: 
        """
        def traverse(node: Node):
            node.label = int(node.label)
            res = np.zeros((2,))
            if node.isLeaf:
                node.word = self.word_map.get(node.word.lower(), 0)
                node.actf = self.L[node.word]
                node.prob = Model.softmax(self.Ws, self.bs, node.actf)
                return np.asarray([np.argmax(node.prob) == node.label, 1])
            else:
                res += traverse(node.left)
                res += traverse(node.right)
                self.activate_node(node, node.left.actf, node.right.actf)
                # print("node.prob: {p} - node.label: {l}.".format(p=node.prob, l=node.label))
                return res + np.asarray([np.argmax(node.prob) == node.label, 1])
        result = traverse(tree.root)
        return np.concatenate((result, np.asarray([np.argmax(tree.root.prob) == tree.root.label, 1])))

    # endregion
